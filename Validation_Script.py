from os import path
import openpyxl
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import random
import smtplib
import itertools
from email.message import EmailMessage
import os

root = tk.Tk()
root.withdraw()
# opening file#####
file_path = filedialog.askopenfilename()
work_book = openpyxl.load_workbook(file_path,keep_vba=True, read_only=False)
work_sheet=work_book["Team Data"]
work_sheet2=work_book["Verification"]
row_count=work_sheet.max_row
row_column=work_sheet.max_column

## dashboard having corresponded_by with sample size #####
xl = pd.ExcelFile(file_path) # or whatever your filename is
df = xl.parse("Team Data", header=None,skiprows=1)
answer = df[4].value_counts()
print(answer)

# work_sheet.cell(row=1,column=row_column).value="S. No."
count=1
max_row_for_sno = max((c.row for c in work_sheet['A'] if c.value is not None))
# print(max_row_for_sno)
s=max_row_for_sno+1
max_column_for_sno = max((c.column for c in work_sheet['1'] if c.value is not None))
p=max_column_for_sno+1
work_sheet.cell(row=1,column=p).value="S.No"
for i in range(2,s):
    work_sheet.cell(row=i,column=p).value=count
    count=count+1

# ## Randomly case id selection####
def caseIdSelection(count,auditor_name):
    case_id=[]
    for i in range(2,row_count+1):
        if work_sheet.cell(row=i,column=5).value==auditor_name:
            case_id.append(work_sheet.cell(row=i,column=p).value)
    res=random.sample(case_id, int(count))    
    return (res)
    
corresponded_by=[]
corresponded_by.extend(df[4].unique())
# print(corresponded_by)

d={}
all_selected_case_id=[]
# choosing samples with auditor######
def Backend_process(count,auditor_name,total_count):
    id=caseIdSelection(count,auditor_name)
    total_count=total_count+int(count)
    key=auditor_name
    value=id
    d[key]=value
    all_selected_case_id.append(id)
    return total_count
# try:
total_count = 0
exitline = True
while True:
    if exitline == True: 
        print(['Enter E for Exit'])
        exitline = False
    Input_value=input("Assign Samples with Auditor: ")
    if len(Input_value.strip()) == 1 and 'E' in Input_value: break
    Input_value = Input_value.split(" ")
    count,auditor_name=Input_value[0],Input_value[1]
    total_count=Backend_process(count,auditor_name,total_count)
# print(d)
merged = list(itertools.chain(*all_selected_case_id))
# print(merged)
### for copying row 1 data of whole column from team data to verification sheet
for k in range(1, 15):
    work_sheet2.cell(row=1,column=k).value = work_sheet.cell(row=1,column=k).value

for i in range(2, row_count+1):
    for j in merged:
        if work_sheet.cell(row=i,column=p).value==j:
            for k in range(1, 15):
                work_sheet2.cell(row=i,column=k).value = work_sheet.cell(row=i,column=k).value

# print("Data copied")
index_row = []
# loop each row in column A
verification_row_count=work_sheet2.max_row
for i in range(1, verification_row_count):
    # define emptiness of cell
    if work_sheet2.cell(i, 1).value is None:
        # collect indexes of rows
        index_row.append(i)

# loop each index value
for row_del in range(len(index_row)):
    work_sheet2.delete_rows(idx=index_row[row_del], amount=1)
    # exclude offset of rows through each iteration
    index_row = list(map(lambda k: k - 1, index_row))


def validation(caseid,lst_selected_By_verifier):
    for i in range(2,row_count+1):
        for j in caseid:
            if j==work_sheet.cell(row=i,column=p).value:
                lst_selected_By_verifier.append(work_sheet.cell(row=i,column=1).value)
    return lst_selected_By_verifier
def comprehension(a, b):
    return [x for x in a if x not in b]

totalst=merged
verifiers=[]

def Selection(number,verifier_id,totalst,lst_selected_By_verifier):
    caseid=random.sample(totalst, int(number))
    lst_selected_By_verifier=validation(caseid,lst_selected_By_verifier)
    # print(lst_selected_By_verifier)
    for i in range(2,verification_row_count+1):
        for j in lst_selected_By_verifier:
            if work_sheet2.cell(row=i,column=1).value==j:
                work_sheet2.cell(row=i,column=15).value=verifier_id
    # print(totalst)
    # print(caseid)
    restcase_id=comprehension(totalst,caseid)
    # print(restcase_id) 
    return restcase_id

print("Total Samples: "+str(total_count))
while True:
    lst_selected_By_verifier=[]
    verifier=input("Assign Verifier: ")
    if len(verifier.strip()) == 1 and 'E' in verifier: break
    verifier = verifier.split(" ")
    number,verifier_id=verifier[0],verifier[1]
    print("Total sample picked : "+number+" out of "+str(total_count))
    totalst=Selection(number,verifier_id,totalst,lst_selected_By_verifier)
    total_count=total_count-int(number)
    verifiers.append(verifier_id)

print("wait we are processing")
work_sheet2.cell(row=1,column=15).value="Verifier"    
work_book.save(file_path)

# ### Sending mail ######
try:
    SENDER_EMAIL = "ayushisahuas24@gmail.com"
    APP_PASSWORD = "XXXXXXX"

    def send_mail_with_excel(recipient_email, subject, content, excel_file):
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email
        msg.set_content(content)
        filename=os.path.basename(excel_file)
        with open(excel_file, 'rb') as f:
            file_data = f.read()
        msg.add_attachment(file_data, maintype="application", subtype="xlsm", filename=filename)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.send_message(msg)

    for i in verifiers:
        # s=i+"@amazon.com"
        s=i+"@gmail.com"
        recipient_email=s
        subject="Test Mail"
        content="Hello, This is system test"
        excel_file=file_path
        send_mail_with_excel(recipient_email, subject, content, excel_file)
except:
    print("Unable to send mail!! please try again")
# recipient_email="sahuayushi74@gmail.com"
# subject="Test Mail"
# content="Hello, This is system test"
# excel_file=file_path
# send_mail_with_excel(recipient_email, subject, content, excel_file)
# except:
#     print("Invalid Input")
